import os
from openpyxl import Workbook
from openpyxl import load_workbook





def _get_caption_order():
    strings = """
# As names should be...
1 - Alias (input)
2 - Application Number
3 - Legal status
4 - Application date
5 - Priority information
6 - Title
7 - Applicants / Proprietors
8 - Pub. No.
9 - Agent
10 - Inventors
11 - Publication Date
12 - Grant Date
13 - Other Title
14 - Parent application
15 - Application Source Type
16 - IPC
17 - Publication Language / Procedural language
18 - Filing Language
19 - Not In Force Date
20 - Grant of Patent (Notification under Section 18(4))
21 - Publication of notice in the Patents and Designs Journal (Section 25(1))
22 - Next Renewal Date
23 - Last annual fee paid
24 - Paid annual fees
25 - Examiner
26 - Original Applicant
27 - Abstract
28 - Attorney Docket Number
29 - Entity Status
30 - AIA (First Inventor to File)
31 - Status Date
32 - International Registration Number (Hague)
33 - International Registration Publication Date
34 - Total PTA Adjustments
35 - Divisional application(s)
36 - Parent Continuity Data
37 - Most recent event
38 - International and Supplementary search report(s)
39 - International and Supplementary search report(s)
40 - Designated contracting states
41 - Lapses during opposition
42 - Opposition
43 - Extension states
44 - US Patent Number
45 - Designated contracting states



1 - Alias
2 - Application Number
2 - Application number:
2 - International Application No.:
3 - Status
3 - Legal stat
4 - Lodged Date
4 - Filing Date
4 - Filing date of the application:
4 - Application date:
4 - International Filing Date:
5 - Priorities Claimed
5 - Priority Claimed
5 - Priority information:
5 - Priority Data:
6 - Application Title
6 - Title in English:
7 - Applicants / Proprietors
7 - Applicant / Proprietor
7 - Applicant/Owner:
7 - Applicant / Owner:
7 - Applicants:
8 - Publication Number
8 - Patent No. / Registration No. :
8 - Pub. No.:
9 - Address for Service
9 - Agent:
10 - Inventors
10 - Inventor:
10 - Inventors:
11 - Publication Date
11 - Publication date:
11 - Publication Date:
12 - Grant Date
12 - Grant date/Registration date:
12 - Grant date / Registration date:
13 - Grant Title
13 - Title in Finnish:
14 - Earlier Application
14 - Parent application:
15 - Application Source
15 - Application type:
15 - Application/Patent type:
15 - Application / Patent type:
16 - IPC classes:
16 - IPC:
17 - Publication Language:
18 - Filing Language:
19 - Not In Force Date
20 - Grant of Patent (Notification under Section 18(4)):
21 - Publication of notice in the Patents and Designs Journal (Section 25(1)):
22 - Next Renewal Date
23 - Last annual fee paid:
24 - Paid annual fees:
25 - Examiner:
26 - Original Applicant:
26 - Original applicant :
27 - Abstract of patent:
27 - Abstract:
28 -
29 -
30 -
31 -
32 -
33 -
34 -
35 - Divisional Application
36 -
37 -
38 -
39 -
40 -
41 -
42 -
43 -
44 -
45 -

#http://patent.prh.fi/patinfo/default2.asp
1 - Alias
2 - Application number:
15 - Application type:
8 - Patent No. / Registration No. :
15 - Application/Patent type:
4 - Filing date of the application:
12 - Grant date/Registration date:
3 - Legal status:
24 - Paid annual fees:
5 - Priority information:
16 - IPC classes:
25 - Examiner:
7 - Applicant/Owner:
10 - Inventor:
9 - Agent:
13 - Title in Finnish:     
11 - Publication date:
23 - Last annual fee paid:
27 - Abstract of patent:
14 - Parent application:
4 - Application date:
8 - Patent No. / Registration No. :
15 - Application / Patent type:
12 - Grant date / Registration date:
7 - Applicant / Owner:
26 - Original Applicant:
6 - Title in English:     


#https://register.epo.org
1 - Alias
8 - About this file
3 - Status
37 - Most recent event
7 - Applicant(s)
10 - Inventor(s)
9 - Representative(s)
2 - Application number, filing date
5 - Priority number, date
18 - Filing language
17 - Procedural language
11 - Publication
38 - International and Supplementary search report(s)
39 - International and Supplementary search report(s)
16 - Classification
40 - Designated contracting states
6 - Title
35 - Divisional application(s)
41 - Lapses during opposition
#42 - Appeal following opposition
42 - Opposition(s)
43 - Extension states
45 - Designated contracting states


# http://portal.uspto.gov/pair/PublicPair
1 - Alias
2 - Application Number:
4 - Filing or 371 (c) Date:
15 - Application Type:
25 - Examiner Name:
28 - Attorney Docket Number:
16 - Class / Subclass:
10 - First Named Inventor:
7 - First Named Applicant:
29 - Entity Status:
30 - AIA (First Inventor to File):
6 - Title of Invention:
3 - Status:
31 - Status Date:
8 - Earliest Publication No:
11 - Earliest Publication Date:
44 - Patent Number:
12 - Issue Date of Patent:
32 - International Registration Number (Hague):
33 - International Registration Publication Date:
34 - Total PTA Adjustments:
35 - Child Continuity Data
36 - Parent Continuity Data
5 - Country |Priority |Priority Date ;
#9 - Name:
#9 - Address:
"""


    max_no = 45

    spl = strings.split("\n")

    result = []
    for _ in range(max_no):
        result.append([])

    current_list = None
    for line in spl:
        if line.strip() == "" or line[0] == "#":
            continue

        lines = line.split("-")

        no = int(lines[0].strip())-1
        current_list = result[no]

        caption = lines[1].strip()
        if caption and caption not in current_list:
            current_list.append(caption)


    return result



class Saver():

    CAPTION_ORDER = _get_caption_order()

    BIBLIOGRAPHIC_DATA_CAPTIONS = [
        "Alias",
        "Patent Number:",
        "Application Number:",
        "Filing Date:",
        "Entity:",
        "Expiration:",
        "Total Amt Due:",
        "Surcharge Date:",
        "Surchg Amt Due:",
        "Most recent events (up to 7):",
        "Address for fee purposes:",
        "Issue Date:",
        "Title:",
        "Status:",
        "Window Opens:",
        "Fee Amt Due:",
        "Fee Code:",
        "Surcharge Fee Code:",

        "Open Date 4th Year",
        "Surcharge Date 4th Year",
        "Close Date 4th Year",
        "Open Date 8th Year",
        "Surcharge Date 8th Year",
        "Close Date 8th Year",
        "Open Date 12th Year",
        "Surcharge Date 12th Year",
        "Close Date 12th Year",
    ]


    def _get_sheet_or_create(self, title, captions):
        try:
            return self._wb[title]
        except KeyError:
            new_sheet = self._wb.create_sheet()
            new_sheet.title = title
            new_sheet.append(captions)

            return new_sheet


    def _load(self):
        if not os.path.isfile(self._output_file):
            self._wb = Workbook()

            # Main
            self._ws_main =  self._wb.active
            self._ws_main.title = "Main"
            # writing headers
            row = []
            for captions in self.CAPTION_ORDER:
                if len(captions):
                    row.append(captions[0])
                else:
                    row.append("")
            self._ws_main.append(row)

        else:
            self._wb = load_workbook(self._output_file)
            self._ws_main = self._wb["Main"]


        # Documents
        self._ws_documents = self._get_sheet_or_create("Documents",
            ["Alias", "Date", "Document type", "No. of Pages",
                "Category/Procedure", "Link"])

        # Forms Filed
        self._ws_formsfiled = self._ws_documents

        # Citations
        self._ws_citations = self._get_sheet_or_create("Citations",
            ["Alias", "Document type", "Reference", "Category"])

        # Fees due
        self._ws_fees = self._get_sheet_or_create("Fees",
            ["Alias", "Input", "Due Date", "The sum of [EUR]",
            "Payment Type", "Payment reference", "Account Number",
            "Payment In order No.", "Customer reference"])

        # Event history
        self._ws_evt_history = self._get_sheet_or_create("Event history",
            ["Alias", "Input", "Date", "Action",
                            "European Patent Bulletin date", "Issue number",
                            "Category"])

        # pto event
        self._ws_pto_event = self._get_sheet_or_create("PTO EVENT",
            ["Input", "Alias", "Viraston päätöksen lähettämispvm",
                "Hakijan määräaika vastata", "Hakijan vastauspvm",
                "Kirjeen nimi"])

        # National Phase
        self._ws_national_phase = self._get_sheet_or_create("FAMILY MEMBER",
            ["Input", "Alias", "Office", "Entry Date", "National Number",
                "National Status"])

        # Patent Family
        self._ws_patent_family = self._get_sheet_or_create("Patent family",
            ["Alias",
                "Source", "Publication", "Publication_date", "Publication_type",
                "Priority No.", "Priority_date"])

        # Bibliographic Data
        self._ws_bibliographic_data = self._get_sheet_or_create(
            "Bibliographic Data", self.BIBLIOGRAPHIC_DATA_CAPTIONS)

        self._ws_parent_continuity = self._get_sheet_or_create(
            "Parent continuity",
            ["Alias",
               'Description', 'Parent Number', 'Parent Filing or 371(c) Date',
               'AIA(First Inventor to File)', 'Parent Status', 'Patent Number'])

        self._ws_design_data = self._get_sheet_or_create(
            "Designated states",
            ["Alias", 'Country Code', 'Lapse', 'Link',])

        # saving
        self.flush()


    def cancel_changes(self):
        self._load()


    def __init__(self, output_filename="Output/data.xlsx", autoflush=False):
        file_dir = os.path.dirname(os.path.realpath(__file__))
        self._output_file = os.path.join(file_dir, output_filename)
        self._do_autoflush = autoflush
        self._load()




    def flush(self):
        self._wb.save(self._output_file)



    def _autoflush(self):
        if self._do_autoflush:
            self.flush()


    def _data_to_row(self, dict_data, caption_order):
        row = []

        for caption in caption_order:

            found = False

            if isinstance(caption, list):
                for cap in caption:
                    if cap in dict_data and dict_data[cap]:
                        row.append(dict_data[cap])
                        found = True
                        break
            else:
                if caption in dict_data:
                    row.append(dict_data[caption])
                    found = True

            if not found:
                row.append("")

        return row



    def save_main_data(self, main_data):
        # self._ws_main.append(self._prepare_main_data_row(main_data))
        self._ws_main.append(self._data_to_row(main_data, self.CAPTION_ORDER))
        self._autoflush()



    def save_formfields_data(self, data):
        caption_order = ("Alias", "Date", "Description")

        for d in data:
            self._ws_formsfiled.append(self._data_to_row(d, caption_order))

        self._autoflush()



    def save_citations_data(self, data):
        caption_order = ("Alias", "Document Type", "Reference", "Category")

        for d in data:
            self._ws_citations.append(self._data_to_row(d, caption_order))

        self._autoflush()


    def save_documents_data(self, data):
        caption_order = ("Alias", "Date", "Document type", "Number of pages",
            "Category/Procedure", "Link")

        if "Category" in data:
            data["Category/Procedure"] = data["Category"]
        if "Procedure" in data:
            data["Category/Procedure"] = data["Procedure"]

        for d in data:
            self._ws_documents.append(self._data_to_row(d, caption_order))

        self._autoflush()



    def save_fees_data(self, data):
        caption_order = ("Alias", "Input", "Due Date", "The sum of [EUR]",
                        "Payment Type", "Payment reference", "Account Number",
                        "Payment In order No.", "Customer reference")
        for d in data:
            self._ws_fees.append(self._data_to_row(d, caption_order))

        self._autoflush()


    def save_evt_history_data(self, data):
        caption_order = ("Alias", "Input", "Date", "Action",
                            "European Patent Bulletin date", "Issue number",
                            "Category")

        for d in data:
            self._ws_evt_history.append(self._data_to_row(d, caption_order))

        self._autoflush()


    def save_pto_event_data(self, data):
        caption_order = ("Alias", "Input", "Viraston päätöksen lähettämispvm",
                "Hakijan määräaika vastata", "Hakijan vastauspvm",
                "Kirjeen nimi")

        for d in data:
            self._ws_pto_event.append(self._data_to_row(d, caption_order))

        self._autoflush()


    def save_national_phase(self, data):
        caption_order = ["Input", "Alias", "Office", "Entry Date",
            "National Number", "National Status"]

        for d in data:
            self._ws_national_phase.append(self._data_to_row(d, caption_order))

        self._autoflush()


    def save_patent_family(self, data_list):
        for i in range(len(data_list)):
            data_list[i] = data_list[i].strip()
            if data_list[i] and data_list[i][len(data_list[i])-1] == ';':
                data_list[i] = data_list[i][:-1]

        self._ws_patent_family.append(data_list)

        self._autoflush()



    def save_parent_continuity(self, data):
        caption_order = ["Alias",
               'Description', 'Parent Number', 'Parent Filing or 371(c) Date',
               'AIA(First Inventor to File)', 'Parent Status', 'Patent Number']

        for d in data:
            self._ws_parent_continuity.append(
                                            self._data_to_row(d, caption_order))

        self._autoflush()



    def save_design_data(self, data):
        caption_order = ["Alias", 'Country Code', 'Lapse', 'Link',]

        for d in data:
            self._ws_design_data.append(self._data_to_row(d, caption_order))

        self._autoflush()



    def save_bibliographic_data(self, data):

        self._ws_bibliographic_data.append(
            self._data_to_row(data, self.BIBLIOGRAPHIC_DATA_CAPTIONS))

        self._autoflush()
